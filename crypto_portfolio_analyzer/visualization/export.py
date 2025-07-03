"""Data export and import system for portfolio data."""

import csv
import json
import logging
from datetime import datetime, timezone
from typing import Dict, List, Optional, Any, Union
from enum import Enum
from dataclasses import dataclass, field
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from ..analytics.models import PortfolioSnapshot, PortfolioHolding, PerformanceMetrics
from ..data.models import HistoricalPrice

logger = logging.getLogger(__name__)


class ExportFormat(Enum):
    """Export format enumeration."""
    CSV = "csv"
    JSON = "json"
    EXCEL = "xlsx"
    PARQUET = "parquet"


@dataclass
class ExportConfig:
    """Export configuration settings."""
    format: ExportFormat
    include_headers: bool = True
    include_metadata: bool = True
    date_format: str = "%Y-%m-%d %H:%M:%S"
    decimal_places: int = 6
    output_path: Optional[str] = None
    compress: bool = False
    custom_fields: List[str] = field(default_factory=list)


class BaseExporter:
    """Base class for all exporters."""
    
    def __init__(self, config: ExportConfig):
        """Initialize base exporter.
        
        Args:
            config: Export configuration
        """
        self.config = config
    
    def export(self, data: Any, filename: str) -> str:
        """Export data to file.
        
        Args:
            data: Data to export
            filename: Output filename
            
        Returns:
            Path to exported file
        """
        raise NotImplementedError("Subclasses must implement export method")
    
    def _prepare_output_path(self, filename: str) -> Path:
        """Prepare output file path.
        
        Args:
            filename: Filename
            
        Returns:
            Path object
        """
        if self.config.output_path:
            output_path = Path(self.config.output_path) / filename
        else:
            output_path = Path(filename)
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        return output_path


class CSVExporter(BaseExporter):
    """CSV data exporter."""
    
    def export(self, data: Union[List[PortfolioSnapshot], List[HistoricalPrice]], filename: str) -> str:
        """Export data to CSV file.
        
        Args:
            data: Data to export
            filename: Output filename
            
        Returns:
            Path to exported file
        """
        output_path = self._prepare_output_path(filename)
        
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            if isinstance(data[0], PortfolioSnapshot):
                self._export_portfolio_snapshots_csv(data, csvfile)
            elif isinstance(data[0], HistoricalPrice):
                self._export_historical_prices_csv(data, csvfile)
            else:
                self._export_generic_csv(data, csvfile)
        
        logger.info(f"Data exported to CSV: {output_path}")
        return str(output_path)
    
    def _export_portfolio_snapshots_csv(self, snapshots: List[PortfolioSnapshot], csvfile):
        """Export portfolio snapshots to CSV."""
        fieldnames = [
            'timestamp', 'portfolio_value', 'total_cost', 'total_return',
            'return_percentage', 'holdings_count'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        if self.config.include_headers:
            writer.writeheader()
        
        for snapshot in snapshots:
            total_return = float(snapshot.portfolio_value - snapshot.total_cost)
            return_percentage = (total_return / float(snapshot.total_cost) * 100) if snapshot.total_cost > 0 else 0
            
            writer.writerow({
                'timestamp': snapshot.timestamp.strftime(self.config.date_format),
                'portfolio_value': round(float(snapshot.portfolio_value), self.config.decimal_places),
                'total_cost': round(float(snapshot.total_cost), self.config.decimal_places),
                'total_return': round(total_return, self.config.decimal_places),
                'return_percentage': round(return_percentage, 2),
                'holdings_count': len(snapshot.holdings)
            })
    
    def _export_historical_prices_csv(self, prices: List[HistoricalPrice], csvfile):
        """Export historical prices to CSV."""
        fieldnames = ['timestamp', 'symbol', 'price', 'volume', 'market_cap']
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        if self.config.include_headers:
            writer.writeheader()
        
        for price in prices:
            writer.writerow({
                'timestamp': price.timestamp.strftime(self.config.date_format),
                'symbol': price.symbol,
                'price': round(float(price.price), self.config.decimal_places),
                'volume': round(float(price.volume or 0), self.config.decimal_places),
                'market_cap': round(float(price.market_cap or 0), self.config.decimal_places)
            })
    
    def _export_generic_csv(self, data: List[Dict], csvfile):
        """Export generic data to CSV."""
        if not data:
            return
        
        fieldnames = list(data[0].keys())
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        
        if self.config.include_headers:
            writer.writeheader()
        
        for row in data:
            writer.writerow(row)


class JSONExporter(BaseExporter):
    """JSON data exporter."""
    
    def export(self, data: Any, filename: str) -> str:
        """Export data to JSON file.
        
        Args:
            data: Data to export
            filename: Output filename
            
        Returns:
            Path to exported file
        """
        output_path = self._prepare_output_path(filename)
        
        # Convert data to JSON-serializable format
        json_data = self._prepare_json_data(data)
        
        with open(output_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(json_data, jsonfile, indent=2, default=str, ensure_ascii=False)
        
        logger.info(f"Data exported to JSON: {output_path}")
        return str(output_path)
    
    def _prepare_json_data(self, data: Any) -> Dict[str, Any]:
        """Prepare data for JSON serialization."""
        result = {
            'exported_at': datetime.now(timezone.utc).isoformat(),
            'format_version': '1.0',
            'data': []
        }
        
        if self.config.include_metadata:
            result['metadata'] = {
                'total_records': len(data) if isinstance(data, list) else 1,
                'decimal_places': self.config.decimal_places,
                'date_format': self.config.date_format
            }
        
        if isinstance(data, list) and data:
            if isinstance(data[0], PortfolioSnapshot):
                result['data_type'] = 'portfolio_snapshots'
                result['data'] = self._serialize_portfolio_snapshots(data)
            elif isinstance(data[0], HistoricalPrice):
                result['data_type'] = 'historical_prices'
                result['data'] = self._serialize_historical_prices(data)
            else:
                result['data_type'] = 'generic'
                result['data'] = data
        else:
            result['data'] = data
        
        return result
    
    def _serialize_portfolio_snapshots(self, snapshots: List[PortfolioSnapshot]) -> List[Dict]:
        """Serialize portfolio snapshots."""
        return [
            {
                'timestamp': snapshot.timestamp.isoformat(),
                'portfolio_value': round(float(snapshot.portfolio_value), self.config.decimal_places),
                'total_cost': round(float(snapshot.total_cost), self.config.decimal_places),
                'holdings': [
                    {
                        'symbol': holding.symbol,
                        'amount': round(float(holding.quantity), self.config.decimal_places),
                        'current_value': round(float(holding.market_value), self.config.decimal_places),
                        'cost_basis': round(float(holding.cost_basis), self.config.decimal_places)
                    }
                    for holding in snapshot.holdings
                ]
            }
            for snapshot in snapshots
        ]
    
    def _serialize_historical_prices(self, prices: List[HistoricalPrice]) -> List[Dict]:
        """Serialize historical prices."""
        return [
            {
                'timestamp': price.timestamp.isoformat(),
                'symbol': price.symbol,
                'price': round(float(price.price), self.config.decimal_places),
                'volume': round(float(price.volume or 0), self.config.decimal_places),
                'market_cap': round(float(price.market_cap or 0), self.config.decimal_places)
            }
            for price in prices
        ]


class ExcelExporter(BaseExporter):
    """Excel data exporter with formatting."""
    
    def export(self, data: Any, filename: str) -> str:
        """Export data to Excel file.
        
        Args:
            data: Data to export
            filename: Output filename
            
        Returns:
            Path to exported file
        """
        output_path = self._prepare_output_path(filename)
        
        # Create workbook
        workbook = openpyxl.Workbook()
        
        if isinstance(data, list) and data:
            if isinstance(data[0], PortfolioSnapshot):
                self._export_portfolio_snapshots_excel(data, workbook)
            elif isinstance(data[0], HistoricalPrice):
                self._export_historical_prices_excel(data, workbook)
            else:
                self._export_generic_excel(data, workbook)
        
        # Remove default sheet if we created others
        if len(workbook.worksheets) > 1:
            workbook.remove(workbook.worksheets[0])
        
        workbook.save(output_path)
        logger.info(f"Data exported to Excel: {output_path}")
        return str(output_path)
    
    def _export_portfolio_snapshots_excel(self, snapshots: List[PortfolioSnapshot], workbook):
        """Export portfolio snapshots to Excel."""
        # Portfolio summary sheet
        ws_summary = workbook.create_sheet("Portfolio Summary")
        
        # Headers
        headers = ['Timestamp', 'Portfolio Value', 'Total Cost', 'Total Return', 'Return %', 'Holdings Count']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, snapshot in enumerate(snapshots, 2):
            total_return = float(snapshot.portfolio_value - snapshot.total_cost)
            return_percentage = (total_return / float(snapshot.total_cost) * 100) if snapshot.total_cost > 0 else 0
            
            ws_summary.cell(row=row, column=1, value=snapshot.timestamp.strftime(self.config.date_format))
            ws_summary.cell(row=row, column=2, value=round(float(snapshot.portfolio_value), 2))
            ws_summary.cell(row=row, column=3, value=round(float(snapshot.total_cost), 2))
            ws_summary.cell(row=row, column=4, value=round(total_return, 2))
            ws_summary.cell(row=row, column=5, value=round(return_percentage, 2))
            ws_summary.cell(row=row, column=6, value=len(snapshot.holdings))
        
        # Auto-adjust column widths
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        # Holdings detail sheet (latest snapshot)
        if snapshots:
            latest_snapshot = snapshots[-1]
            ws_holdings = workbook.create_sheet("Current Holdings")
            
            holdings_headers = ['Symbol', 'Amount', 'Current Value', 'Cost Basis', 'P&L', 'P&L %']
            for col, header in enumerate(holdings_headers, 1):
                cell = ws_holdings.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row, holding in enumerate(latest_snapshot.holdings, 2):
                pnl = float(holding.market_value - holding.cost_basis)
                pnl_percent = (pnl / float(holding.cost_basis) * 100) if holding.cost_basis > 0 else 0

                ws_holdings.cell(row=row, column=1, value=holding.symbol)
                ws_holdings.cell(row=row, column=2, value=round(float(holding.quantity), self.config.decimal_places))
                ws_holdings.cell(row=row, column=3, value=round(float(holding.market_value), 2))
                ws_holdings.cell(row=row, column=4, value=round(float(holding.cost_basis), 2))
                ws_holdings.cell(row=row, column=5, value=round(pnl, 2))
                ws_holdings.cell(row=row, column=6, value=round(pnl_percent, 2))
            
            # Auto-adjust column widths
            for column in ws_holdings.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_holdings.column_dimensions[column_letter].width = adjusted_width
    
    def _export_historical_prices_excel(self, prices: List[HistoricalPrice], workbook):
        """Export historical prices to Excel."""
        ws = workbook.create_sheet("Historical Prices")
        
        headers = ['Timestamp', 'Symbol', 'Price', 'Volume', 'Market Cap']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for row, price in enumerate(prices, 2):
            ws.cell(row=row, column=1, value=price.timestamp.strftime(self.config.date_format))
            ws.cell(row=row, column=2, value=price.symbol)
            ws.cell(row=row, column=3, value=round(float(price.price), self.config.decimal_places))
            ws.cell(row=row, column=4, value=round(float(price.volume or 0), self.config.decimal_places))
            ws.cell(row=row, column=5, value=round(float(price.market_cap or 0), self.config.decimal_places))
    
    def _export_generic_excel(self, data: List[Dict], workbook):
        """Export generic data to Excel."""
        if not data:
            return
        
        ws = workbook.create_sheet("Data")
        
        # Convert to DataFrame for easier handling
        df = pd.DataFrame(data)
        
        # Write to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")


class DataExporter:
    """Main data export manager."""
    
    def __init__(self):
        """Initialize data exporter."""
        self.exporters = {
            ExportFormat.CSV: CSVExporter,
            ExportFormat.JSON: JSONExporter,
            ExportFormat.EXCEL: ExcelExporter
        }
    
    def export_data(self, data: Any, config: ExportConfig, filename: str) -> str:
        """Export data using specified configuration.
        
        Args:
            data: Data to export
            config: Export configuration
            filename: Output filename
            
        Returns:
            Path to exported file
        """
        if config.format not in self.exporters:
            raise ValueError(f"Unsupported export format: {config.format}")
        
        exporter_class = self.exporters[config.format]
        exporter = exporter_class(config)
        
        return exporter.export(data, filename)
    
    def export_portfolio_snapshots(self, snapshots: List[PortfolioSnapshot], 
                                 format: ExportFormat = ExportFormat.CSV,
                                 filename: Optional[str] = None) -> str:
        """Export portfolio snapshots.
        
        Args:
            snapshots: Portfolio snapshots to export
            format: Export format
            filename: Output filename
            
        Returns:
            Path to exported file
        """
        config = ExportConfig(format=format)
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"portfolio_snapshots_{timestamp}.{format.value}"
        
        return self.export_data(snapshots, config, filename)
    
    def export_historical_prices(self, prices: List[HistoricalPrice],
                               format: ExportFormat = ExportFormat.CSV,
                               filename: Optional[str] = None) -> str:
        """Export historical prices.
        
        Args:
            prices: Historical prices to export
            format: Export format
            filename: Output filename
            
        Returns:
            Path to exported file
        """
        config = ExportConfig(format=format)
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"historical_prices_{timestamp}.{format.value}"
        
        return self.export_data(prices, config, filename)
